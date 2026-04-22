import sqlite3
from openpyxl import load_workbook

DB_NAME = 'obuv.db'

conn = sqlite3.connect(DB_NAME)
cursor = conn.cursor()
cursor.execute("PRAGMA foreign_keys = ON;")
cursor.executescript('''
    CREATE TABLE IF NOT EXISTS categories (id INTEGER PRIMARY KEY, name TEXT);
    CREATE TABLE IF NOT EXISTS manufacturers (id INTEGER PRIMARY KEY, name TEXT);
    CREATE TABLE IF NOT EXISTS suppliers (id INTEGER PRIMARY KEY, name TEXT);
    CREATE TABLE IF NOT EXISTS pickup_points (id INTEGER PRIMARY KEY, address TEXT);
    CREATE TABLE IF NOT EXISTS users (id INTEGER PRIMARY KEY, role, full_name, login, password);
    CREATE TABLE IF NOT EXISTS products (
        id INTEGER PRIMARY KEY, sku, name, unit_of_measurement, price, supplier_id,
        manufacturer_id, category_id, discount, stock_quantity, description, photo_filename,
        FOREIGN KEY (supplier_id) REFERENCES suppliers(id),
        FOREIGN KEY (manufacturer_id) REFERENCES manufacturers(id),
        FOREIGN KEY (category_id) REFERENCES categories(id)
    );
    CREATE TABLE IF NOT EXISTS orders (
        id INTEGER PRIMARY KEY, product_id, quantity, order_date, issue_date, pickup_point_id,
        user_full_name, pickup_code, status,
        FOREIGN KEY (product_id) REFERENCES products(sku) ON DELETE SET NULL ON UPDATE CASCADE,
        FOREIGN KEY (pickup_point_id) REFERENCES pickup_points(id) ON DELETE SET NULL
    );
''')
conn.commit()
conn.close()

conn = sqlite3.connect(DB_NAME)
cursor = conn.cursor()

# --- tovar_import ---
try:
    wb = load_workbook('data/tovar_import.xlsx')
    ws = wb.active
    if ws:
        categories_set = set()
        manufacturers_set = set()
        suppliers_set = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 11: continue
            sku, name, unit, price, sup_id, man_id, cat_id, disc, stock, desc, photo = row[:11]

            # ✅ Преобразуем в int через str
            try: cat_id_int = int(str(cat_id)) if cat_id is not None and str(cat_id).strip() != '' else None
            except: cat_id_int = None
            try: man_id_int = int(str(man_id)) if man_id is not None and str(man_id).strip() != '' else None
            except: man_id_int = None
            try: sup_id_int = int(str(sup_id)) if sup_id is not None and str(sup_id).strip() != '' else None
            except: sup_id_int = None

            if cat_id_int:
                name_map = {1: "Женская обувь", 2: "Мужская обувь"}
                categories_set.add((cat_id_int, name_map.get(cat_id_int, f'Категория {cat_id_int}')))

            if man_id_int:
                name_map = {1: "Kari", 2: "Marco Tozzi", 3: "Рос", 4: "Rieker", 5: "Alessio Nesca", 6: "CROSBY"}
                manufacturers_set.add((man_id_int, name_map.get(man_id_int, f'Производитель {man_id_int}')))

            if sup_id_int:
                name_map = {1: "Kari", 2: "Обувь для вас"}
                suppliers_set.add((sup_id_int, name_map.get(sup_id_int, f'Поставщик {sup_id_int}')))

            cursor.execute(
                "INSERT OR IGNORE INTO products (sku, name, unit_of_measurement, price, supplier_id, manufacturer_id, category_id, discount, stock_quantity, description, photo_filename) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (sku, name, unit, price or 0, sup_id_int, man_id_int, cat_id_int, disc or 0, stock or 0, desc, photo)
            )

        cursor.executemany("INSERT OR IGNORE INTO categories (name) VALUES (?)", [(name,) for _, name in sorted(categories_set)])
        cursor.executemany("INSERT OR IGNORE INTO manufacturers (name) VALUES (?)", [(name,) for _, name in sorted(manufacturers_set)])
        cursor.executemany("INSERT OR IGNORE INTO suppliers (name) VALUES (?)", [(name,) for _, name in sorted(suppliers_set)])
    else:
        print("Лист tovar_import.xlsx пуст.")
except FileNotFoundError:
    print("Файл tovar_import.xlsx не найден.")

# --- user_import ---
try:
    wb = load_workbook('data/user_import.xlsx')
    ws = wb.active
    if ws:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 4: continue
            role, full_name, login, password = row[:4]
            if all([role, full_name, login, password]):
                cursor.execute("INSERT OR IGNORE INTO users (role, full_name, login, password) VALUES (?, ?, ?, ?)", (role, full_name, login, password))
    else:
        print("Лист user_import.xlsx пуст.")
except FileNotFoundError:
    print("Файл user_import.xlsx не найден.")

# --- punkt_import ---
try:
    wb = load_workbook('data/punkt_import.xlsx')
    ws = wb.active
    if ws:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]: cursor.execute("INSERT OR IGNORE INTO pickup_points (address) VALUES (?)", (row[0],))
    else:
        print("Лист punkt_import.xlsx пуст.")
except FileNotFoundError:
    print("Файл punkt_import.xlsx не найден.")

# --- zakaz_import ---
try:
    wb = load_workbook('data/zakaz_import.xlsx')
    ws = wb.active
    if ws:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 8: continue
            pid, qty, odate, idate, ppid, uname, pcode, stat = row[:8]

            cursor.execute(
                "INSERT OR IGNORE INTO orders (product_id, quantity, order_date, issue_date, pickup_point_id, user_full_name, pickup_code, status) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (pid, qty, odate, idate, ppid, uname, pcode, stat)
            )
    else:
        print("Лист zakaz_import.xlsx пуст.")
except FileNotFoundError:
    print("Файл zakaz_import.xlsx не найден.")

conn.commit()
conn.close()
print("База данных успешно создана и данные импортированы.")